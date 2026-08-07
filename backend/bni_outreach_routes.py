"""
BNI Outreach — Drawlead OS

Tracks prospective companies/individuals being reached out to for BNI chapter
membership: name, brand, chapter of interest, contact details, and an
outreach status pipeline (To do / Contacted / Interested / Not Interested /
Converted).

Storage: collection `bni_outreach`.
"""
import csv
import io
import re
import uuid
import httpx
from fastapi import APIRouter, HTTPException, Request
from pydantic import BaseModel
from typing import Optional
from datetime import datetime, timezone

bni_outreach_router = APIRouter(prefix="/bni/outreach", tags=["bni"])
bni_outreach_sources_router = APIRouter(prefix="/bni/outreach-sources", tags=["bni"])

OUTREACH_STATUSES = [
    "To do", "New Lead", "Contacted", "RNR", "Scheduled One to One",
    "One to One Completed", "Not Interested", "Relationship", "Lead", "Later",
]

# Column synonyms used when importing outreach rows from a Google Sheet source
# (same columns as the CSV import on the Outreach tab).
SOURCE_FIELD_SYNONYMS = {
    "name": ["name"],
    "brand_name": ["brand name", "brand"],
    "chapter_name": ["chapter name", "chapter"],
    "email": ["email", "email address"],
    "profile_link": ["profile link", "profile"],
    "phone": ["phone", "phone number", "mobile", "contact number", "phone 01", "phone1", "phone 1"],
    "phone2": ["phone 02", "phone2", "phone 2", "alternate phone", "secondary phone", "mobile 2", "mobile 02"],
    "website": ["website", "site"],
    "status": ["status"],
    "location": ["location"],
    "category": ["category", "category name", "business category"],
}


def _norm(s):
    return " ".join(str(s or "").lower().split())


def _sheet_id(url: str):
    m = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", url or "")
    return m.group(1) if m else None


def _sheet_csv_url(url: str):
    """Turn a shared Google Sheets URL into its CSV export URL (works when the
    sheet is shared 'anyone with the link can view')."""
    sheet_id = _sheet_id(url)
    if not sheet_id:
        return None
    export = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv"
    gm = re.search(r"[#&?]gid=(\d+)", url or "")
    if gm:
        export += f"&gid={gm.group(1)}"
    return export


def _map_headers(header_row):
    m = {}
    for i, h in enumerate(header_row):
        nh = _norm(h)
        for field, syns in SOURCE_FIELD_SYNONYMS.items():
            if nh in [_norm(x) for x in syns]:
                m[i] = field
                break
    return m


async def _upsert_outreach_row(db, source_id, source_name, name, rec, cat_id, cat_name, group, user_id, now):
    """Insert-or-update one outreach row from a synced source. Deduped by
    (source, category, name). Returns (imported, updated) as 0/1."""
    # Untouched leads from a sheet land in the "New Lead" stage by default.
    status = rec.get("status") if rec.get("status") in OUTREACH_STATUSES else "New Lead"
    # Sheet-owned columns that a re-sync always refreshes. Status and remarks
    # are intentionally excluded on update so a user's manual pipeline changes
    # aren't wiped by the next sync.
    fields = {
        "name": name,
        "brand_name": rec.get("brand_name", ""),
        "chapter_name": rec.get("chapter_name", ""),
        "email": rec.get("email", ""),
        "profile_link": rec.get("profile_link", ""),
        "phone": rec.get("phone", ""),
        "phone2": rec.get("phone2", ""),
        "website": rec.get("website", ""),
        "location": rec.get("location", ""),
        "category_id": cat_id,
        "category_name": cat_name,
        "group": group,
        "source_id": source_id,
        "source_name": source_name,
        "updated_at": now,
    }
    existing = await db.bni_outreach.find_one(
        {
            "source_id": source_id,
            "category_name": cat_name,
            "name": {"$regex": f"^{re.escape(name)}$", "$options": "i"},
        },
        {"_id": 0, "outreach_id": 1},
    )
    if existing:
        await db.bni_outreach.update_one({"outreach_id": existing["outreach_id"]}, {"$set": fields})
        return 0, 1
    fields.update({
        "status": status,
        "remarks": "",
        "outreach_id": f"bniout_{uuid.uuid4().hex[:10]}",
        "created_by": user_id,
        "created_at": now,
    })
    await db.bni_outreach.insert_one(fields)
    return 1, 0


async def _lead_from_outreach(db, entry, user_id):
    """When an outreach prospect is marked 'Lead', push it to the main Leads
    pipeline (source 'BNI'), once per entry."""
    if entry.get("lead_created"):
        return
    stage = await db.lead_stages.find_one(
        {"name": {"$regex": "^lead$", "$options": "i"}, "is_deleted": {"$ne": True}},
        {"_id": 0, "stage_id": 1},
    )
    now = datetime.now(timezone.utc)
    await db.leads_v2.insert_one({
        "lead_id": f"lead_{uuid.uuid4().hex[:12]}",
        "name": entry.get("name", ""),
        "phone": entry.get("phone", ""),
        "email": entry.get("email", ""),
        "company_name": entry.get("brand_name", ""),
        "location": entry.get("location", ""),
        "website": entry.get("website", ""),
        "social_media": entry.get("profile_link", ""),
        "what_do_you_do": "",
        "source": "BNI",
        "lead_owner": user_id,
        "lead_type": "",
        "priority": "Medium",
        "date_of_lead": now.date().isoformat(),
        "industry": entry.get("category_name", ""),
        "estimation": 0,
        "quotation_link": "",
        "proposal_link": "",
        "notes": f"Auto-created from BNI Outreach ({entry.get('chapter_name', '') or 'cross chapter'})",
        "stage_id": stage.get("stage_id") if stage else "",
        "custom_fields": {},
        "pipeline": "pre_sales",
        "created_by": user_id,
        "created_at": now,
        "updated_at": now,
        "is_deleted": False,
    })
    await db.bni_outreach.update_one({"outreach_id": entry["outreach_id"]}, {"$set": {"lead_created": True}})


async def _sync_cross_chapter_oto(db, entry, meeting_status, user_id):
    """Mirror an outreach prospect into the BNI One-to-One tab as a
    cross-chapter one-to-one (kept in sync by outreach_id)."""
    now = datetime.now(timezone.utc).isoformat()
    fields = {
        "member_name": entry.get("name", ""),
        "member_phone": entry.get("phone", ""),
        "member_email": entry.get("email", ""),
        "member_company": entry.get("brand_name", ""),
        "chapter_name": entry.get("chapter_name", "") or "Cross Chapter",
        "meeting_status": meeting_status,
        "updated_at": now,
    }
    existing = await db.bni_one_to_ones.find_one({"outreach_id": entry["outreach_id"]}, {"_id": 0, "entry_id": 1})
    if existing:
        await db.bni_one_to_ones.update_one({"entry_id": existing["entry_id"]}, {"$set": fields})
        return
    fields.update({
        "entry_id": f"bniotoo_{uuid.uuid4().hex[:10]}",
        "entry_type": "cross_chapter",
        "member_id": "",
        "meeting_date": now[:10],
        "meeting_time": "",
        "location": "",
        "invited_by": "me",
        "remark": "",
        "gives": [],
        "referrals": [],
        "status": "",
        "expense_entry_id": "",
        "expense_amount": 0,
        "meeting_mode": "offline",
        "lead_created": False,
        "outreach_id": entry["outreach_id"],
        "created_by": user_id,
        "created_at": now,
    })
    await db.bni_one_to_ones.insert_one(fields)


class OutreachCreate(BaseModel):
    name: str
    brand_name: str = ""
    chapter_name: str = ""
    email: str = ""
    profile_link: str = ""
    phone: str = ""
    phone2: str = ""
    website: str = ""
    status: str = "To do"
    location: str = ""
    category_id: str = ""
    remarks: str = ""


class OutreachUpdate(BaseModel):
    name: Optional[str] = None
    brand_name: Optional[str] = None
    chapter_name: Optional[str] = None
    email: Optional[str] = None
    profile_link: Optional[str] = None
    phone: Optional[str] = None
    phone2: Optional[str] = None
    website: Optional[str] = None
    status: Optional[str] = None
    location: Optional[str] = None
    category_id: Optional[str] = None
    remarks: Optional[str] = None


@bni_outreach_router.get("")
async def list_outreach(request: Request):
    from server import get_current_user, db
    await get_current_user(request)
    return await db.bni_outreach.find({}, {"_id": 0}).sort("created_at", -1).to_list(5000)


@bni_outreach_router.post("")
async def create_outreach(payload: OutreachCreate, request: Request):
    from server import get_current_user, db
    user = await get_current_user(request)
    if not payload.name.strip():
        raise HTTPException(status_code=400, detail="Name is required")

    category_name = ""
    group = ""
    if payload.category_id:
        category = await db.bni_categories.find_one({"category_id": payload.category_id}, {"_id": 0, "name": 1, "group": 1})
        if category:
            category_name = category.get("name", "")
            group = category.get("group") or _group_from(category_name)

    status = payload.status if payload.status in OUTREACH_STATUSES else "To do"
    now = datetime.now(timezone.utc).isoformat()
    doc = {
        "outreach_id": f"bniout_{uuid.uuid4().hex[:10]}",
        "name": payload.name.strip(),
        "brand_name": payload.brand_name.strip(),
        "chapter_name": payload.chapter_name.strip(),
        "email": payload.email.strip(),
        "profile_link": payload.profile_link.strip(),
        "phone": payload.phone.strip(),
        "phone2": payload.phone2.strip(),
        "website": payload.website.strip(),
        "status": status,
        "location": payload.location.strip(),
        "category_id": payload.category_id,
        "category_name": category_name,
        "group": group,
        "remarks": (payload.remarks or "").strip(),
        "created_by": user.user_id,
        "created_at": now,
        "updated_at": now,
    }
    await db.bni_outreach.insert_one(doc)
    doc.pop("_id", None)
    return doc


@bni_outreach_router.put("/{outreach_id}")
async def update_outreach(outreach_id: str, payload: OutreachUpdate, request: Request):
    from server import get_current_user, db
    user = await get_current_user(request)
    entry = await db.bni_outreach.find_one({"outreach_id": outreach_id}, {"_id": 0})
    if not entry:
        raise HTTPException(status_code=404, detail="Outreach entry not found")

    update_data = {k: v for k, v in payload.dict().items() if v is not None}
    if "status" in update_data and update_data["status"] not in OUTREACH_STATUSES:
        raise HTTPException(status_code=400, detail=f"Invalid status: {update_data['status']}")
    if "category_id" in update_data:
        category_name = ""
        group = ""
        if update_data["category_id"]:
            category = await db.bni_categories.find_one({"category_id": update_data["category_id"]}, {"_id": 0, "name": 1, "group": 1})
            if category:
                category_name = category.get("name", "")
                group = category.get("group") or _group_from(category_name)
        update_data["category_name"] = category_name
        update_data["group"] = group
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()

    await db.bni_outreach.update_one({"outreach_id": outreach_id}, {"$set": update_data})
    updated = await db.bni_outreach.find_one({"outreach_id": outreach_id}, {"_id": 0})

    # Status-driven downstream actions.
    new_status = update_data.get("status")
    if new_status == "Lead":
        await _lead_from_outreach(db, updated, user.user_id)
    elif new_status == "Scheduled One to One":
        await _sync_cross_chapter_oto(db, updated, "Scheduled", user.user_id)
    elif new_status == "One to One Completed":
        await _sync_cross_chapter_oto(db, updated, "Completed", user.user_id)

    return await db.bni_outreach.find_one({"outreach_id": outreach_id}, {"_id": 0})


@bni_outreach_router.delete("/{outreach_id}")
async def delete_outreach(outreach_id: str, request: Request):
    from server import get_current_user, db
    await get_current_user(request)
    result = await db.bni_outreach.delete_one({"outreach_id": outreach_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Outreach entry not found")
    return {"success": True}


# ---------- Sources (Google Sheets that sync into the outreach list) ----------

class OutreachSourceCreate(BaseModel):
    name: str
    sheet_url: str


@bni_outreach_sources_router.get("")
async def list_outreach_sources(request: Request):
    from server import get_current_user, db
    await get_current_user(request)
    return await db.bni_outreach_sources.find({}, {"_id": 0}).sort("created_at", -1).to_list(500)


@bni_outreach_sources_router.post("")
async def create_outreach_source(payload: OutreachSourceCreate, request: Request):
    from server import get_current_user, db
    user = await get_current_user(request)
    name = (payload.name or "").strip()
    sheet_url = (payload.sheet_url or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Source name is required")
    if not _sheet_csv_url(sheet_url):
        raise HTTPException(status_code=400, detail="Enter a valid Google Sheets link")
    now = datetime.now(timezone.utc).isoformat()
    doc = {
        "source_id": f"bnisrc_{uuid.uuid4().hex[:10]}",
        "name": name,
        "sheet_url": sheet_url,
        "last_synced_at": None,
        "last_row_count": 0,
        "created_by": user.user_id,
        "created_at": now,
    }
    await db.bni_outreach_sources.insert_one(doc)
    doc.pop("_id", None)
    return doc


@bni_outreach_sources_router.delete("/{source_id}")
async def delete_outreach_source(source_id: str, request: Request):
    from server import get_current_user, db
    await get_current_user(request)
    result = await db.bni_outreach_sources.delete_one({"source_id": source_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Source not found")
    # The outreach rows this source imported are left in place (they may have
    # been edited) — clear their source tag so they read as standalone entries.
    await db.bni_outreach.update_many({"source_id": source_id}, {"$set": {"source_id": "", "source_name": ""}})
    return {"success": True}


def _cell_str(v):
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _extract_row(row, header_map):
    rec = {}
    for i, field in header_map.items():
        rec[field] = _cell_str(row[i]) if i < len(row) else ""
    return rec


def _group_from(name):
    return name.split("(", 1)[0].strip() if "(" in (name or "") else (name or "")


async def _fetch_workbook_sheets(client, sheet_id):
    """Read every worksheet of a public sheet via its XLSX export — reliable
    for 'anyone with the link can view' sheets. Returns [(name, rows)] where
    rows is a list of value-tuples (first row = header), or None if the sheet
    couldn't be read as a workbook (e.g. private, so an HTML login page comes
    back instead of xlsx). Returns None (→ CSV fallback) if openpyxl isn't
    installed yet, so a missing dep degrades gracefully."""
    try:
        import openpyxl
    except Exception:
        return None
    try:
        resp = await client.get(f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx")
    except Exception:
        return None
    if resp.status_code != 200:
        return None
    try:
        wb = openpyxl.load_workbook(io.BytesIO(resp.content), read_only=True, data_only=True)
    except Exception:
        return None
    sheets = []
    for nm in wb.sheetnames:
        ws = wb[nm]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        sheets.append((nm, rows))
    wb.close()
    return sheets


@bni_outreach_sources_router.post("/{source_id}/sync")
async def sync_outreach_source(source_id: str, request: Request):
    """Sync a source Google Sheet into the outreach list. Each worksheet (tab)
    is treated as a category: the tab name is the category name, and the part
    before its first "(" is the group. Falls back to a single-CSV import (with
    a Category column) if the tabs can't be enumerated. Deduped by (source,
    category, name) so re-syncing updates rather than duplicates."""
    from server import get_current_user, db
    user = await get_current_user(request)
    source = await db.bni_outreach_sources.find_one({"source_id": source_id}, {"_id": 0})
    if not source:
        raise HTTPException(status_code=404, detail="Source not found")
    sheet_id = _sheet_id(source.get("sheet_url", ""))
    if not sheet_id:
        raise HTTPException(status_code=400, detail="This source is not a valid Google Sheets link")

    cats = await db.bni_categories.find({}, {"_id": 0, "category_id": 1, "name": 1}).to_list(2000)
    cat_by_name = {_norm(c["name"]): c for c in cats}
    source_name = source.get("name", "")
    now = datetime.now(timezone.utc).isoformat()
    imported = 0
    updated = 0
    total = 0
    tabs = 0

    async with httpx.AsyncClient(follow_redirects=True, timeout=60.0) as client:
        sheets = await _fetch_workbook_sheets(client, sheet_id)

        if sheets is not None:
            # Per-tab: tab name = category, before "(" = group.
            for nm, rows in sheets:
                if not rows:
                    continue
                header_map = _map_headers(rows[0])
                if "name" not in header_map.values():
                    continue
                tabs += 1
                cat = cat_by_name.get(_norm(nm))
                cat_id = cat["category_id"] if cat else ""
                cat_name = nm
                group = _group_from(nm)
                for row in rows[1:]:
                    rec = _extract_row(row, header_map)
                    name = (rec.get("name") or "").strip()
                    if not name:
                        continue
                    total += 1
                    i_add, u_add = await _upsert_outreach_row(db, source_id, source_name, name, rec, cat_id, cat_name, group, user.user_id, now)
                    imported += i_add
                    updated += u_add
        else:
            # Fallback: single CSV sheet, category comes from a Category column.
            csv_url = _sheet_csv_url(source.get("sheet_url", ""))
            try:
                resp = await client.get(csv_url)
            except Exception:
                raise HTTPException(status_code=400, detail="Could not reach the sheet. Check the link and try again.")
            if resp.status_code != 200 or "text/csv" not in (resp.headers.get("content-type") or ""):
                raise HTTPException(status_code=400, detail="Could not read the sheet — share it as 'Anyone with the link can view'.")
            rows = list(csv.reader(io.StringIO(resp.text)))
            if not rows:
                raise HTTPException(status_code=400, detail="The sheet is empty")
            header_map = _map_headers(rows[0])
            if "name" not in header_map.values():
                raise HTTPException(status_code=400, detail="The sheet needs a 'Name' column")
            tabs = 1
            for row in rows[1:]:
                rec = _extract_row(row, header_map)
                name = (rec.get("name") or "").strip()
                if not name:
                    continue
                total += 1
                cat = cat_by_name.get(_norm(rec.get("category", "")))
                cat_id = cat["category_id"] if cat else ""
                cat_name = cat["name"] if cat else rec.get("category", "")
                group = _group_from(cat_name)
                i_add, u_add = await _upsert_outreach_row(db, source_id, source_name, name, rec, cat_id, cat_name, group, user.user_id, now)
                imported += i_add
                updated += u_add

    if tabs == 0 and total == 0:
        raise HTTPException(status_code=400, detail="No rows found — make sure the sheet is shared as 'Anyone with the link can view' and each tab has a 'Name' column.")

    await db.bni_outreach_sources.update_one(
        {"source_id": source_id},
        {"$set": {"last_synced_at": now, "last_row_count": total, "last_tab_count": tabs}},
    )
    return {"imported": imported, "updated": updated, "total_rows": total, "tabs": tabs}
